from flask import Flask, render_template, redirect, request
from time import time, ctime
from uuid import uuid4
from requests import get
from keras.models import load_model
from PIL import Image, ImageOps
import numpy as np
from connector import PostgreSQLConnector
from psycopg2 import connect
from build_tables import build
from users import users
from plants import plants
from openpyxl import load_workbook
from datetime import datetime as dt

path = 'data/device.xlsx'
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active
hour = dt.now().strftime("%H")
minutes = dt.now().strftime("%M")
day = dt.now().strftime("%d")
month = dt.now().strftime("%m")
year = dt.now().strftime("%Y")
app = Flask("__name__")
dbc = PostgreSQLConnector(database = "dbagro", user = "postgres", password = "1234")
build()
users = users(dbc)
plants = plants(dbc)

#TODO: model
models = [load_model("models/apple_model.h5", compile = False), load_model("models/corn_model.h5", compile = False), load_model("models/potato_model.h5", compile = False)]
all_class_names = [open("labels/apple_labels.txt", "r").readlines(), open("labels/corn_labels.txt", "r").readlines(), open("labels/potato_labels.txt", "r").readlines()] 


def detect_class(image_path, plant_index):
    np.set_printoptions(suppress=True)
    data = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)
    #TODO: Replace this with the path to your image
    image = Image.open(image_path).convert("RGB")
    #TODO: resizing the image to be at least 224x224 and then cropping from the center
    size = (224, 224)
    image = ImageOps.fit(image, size, Image.Resampling.LANCZOS)
    #TODO: turn the image into a numpy array
    image_array = np.asarray(image)
    #TODO: Normalize the image
    normalized_image_array = (image_array.astype(np.float32) / 127.5) - 1
    #TODO: Load the image into the array
    data[0] = normalized_image_array
    model = models[plant_index]
    class_names = all_class_names[plant_index]
    #TODO: Predicts the model
    prediction = model.predict(data)
    index = np.argmax(prediction)
    class_name = class_names[index]
    confidence_score = prediction[0][index]
    id = class_name[:2].strip()
    advice = ""
    try:
        with open("Maslahatlar//" + id + ".txt", "r") as file:
            advice = file.read()
    except:
        pass
    return class_name[2:], advice


def get_data_from_device():
    ip = "http://192.168.137.167/"
    date = str(ctime(time()))
    havonamlik = get(ip + "humidity").text
    temperature = get(ip + "temperature").text
    tuproqnamlik = get(ip + "tuproqnamlik").text
    suvbormi = "Bor" if get(ip + "fotorezistor").text == "1" else "Yo'q"
    motoryoniqmi = "O'chiq"
    data = {
        "date": date,
        "havonamlik": havonamlik,
        "temperature": temperature,
        "tuproqnamlik": tuproqnamlik,
        "suvbormi": suvbormi,
        "motoryoniqmi": motoryoniqmi   
    }
    i = sheet_obj.max_row + 1
    A = f"A{i}"
    B = f"B{i}"
    C = f"C{i}"
    D = f"D{i}"
    E = f"E{i}"
    F = f"F{i}"    
    sheet_obj[A].value = date
    sheet_obj[B].value = havonamlik
    sheet_obj[C].value = temperature
    sheet_obj[D].value = tuproqnamlik
    sheet_obj[E].value = suvbormi
    sheet_obj[F].value = motoryoniqmi
    wb_obj.save(path)
    return data


def file_save(request, name):
    file = request.files[name]
    allowed_extensions = []
    if name == "picture":
        allowed_extensions = ["jpg", "jpeg", "png", "bmp"]
    else:
        allowed_extensions = ["docx", "doc", "pdf", "txt"]
    filename = file.filename
    if filename == "":
        return None 
    extension = filename.rsplit(".", 1)[1].lower()
    if extension not in allowed_extensions:
        return None
    save_filepath = "uploads/"+uuid4().hex + f".{extension}"
    file.save("static/" + save_filepath)
    return save_filepath


@app.route("/", methods = ["POST", "GET"])
def index():
    user = users.is_login(request)
    is_login = True
    if not user:
        is_login = False
    is_admin = False
    if is_login:
        if user['login'] == "admin":
            is_admin = True
    return render_template("index.html", is_login = is_login, user = user, is_admin = is_admin)


@app.route("/hudud", methods = ["POST", "GET"])
def hudud():
    user = users.is_login(request)
    is_login = True
    if not user:
        is_login = False
    return render_template("hudud.html", is_login = is_login, user = user)


@app.route("/plant/<id>", methods = ["POST", "GET"])
def plant(id):
    user = users.is_login(request)
    is_login = True
    if not user:
        is_login = False
    is_admin = False
    if is_login:
        if user['login'] == "admin":
            is_admin = True
    if not plants.is_id_valid(id):
        return redirect("/plant/no")
    plant = plants.get_plant(id)
    return render_template("plant.html", plant = plant, user = user, is_login = is_login, is_admin = is_admin)


@app.route("/add_plant", methods = ["POST", "GET"])
def add_plant():
    user = users.is_login(request)
    if not user:
        return redirect("/login")
    if not users.is_admin(user['token']):
        return redirect("/")
    if request.method == "POST":
        data = {
            "name": request.form.get("name"),
            "definition": request.form.get("definition"),
            "pic_path": file_save(request, "picture")
        }
        if not data['pic_path']:
            return redirect("/add_plant?error=0")#rasm fayli formati noto'g'ri
        plants.add(**data)
        return redirect("/add_plant?success=1")#Qo'shildi
    else:
        return render_template("add_plant.html", user = user, is_login = True, is_admin = True)


@app.route("/delete_plant/<id>", methods = ["GET"])
def delete_plant(id):
    user = users.is_login(request)
    if not user:
        return redirect("/login")
    if not users.is_admin(user['token']):
        return redirect("/")
    plants.delete_plant(id)
    return redirect("/plants")


@app.route("/plants", methods = ["POST", "GET"])
def plantss():
    user = users.is_login(request)
    is_login = True
    if not user:
        is_login = False
    is_admin = False
    if is_login:
        if user['login'] == "admin":
            is_admin = True
    all_plants = plants.get_all()
    return render_template("plants.html", is_login = is_login, user = user, plants = all_plants, is_admin = is_admin)


@app.route("/my_device", methods = ["POST", "GET"])
def my_device():
    user = users.is_login(request)
    is_login = True
    if not user:
        return redirect("/login")
    is_admin = False
    if is_login:
        if user['login'] == "admin":
            is_admin = True
    if request.method == "GET":
        return render_template("my_device.html", is_login = is_login, user = user, is_admin = is_admin)
    data = get_data_from_device()
    return data


@app.route("/detect_disease", methods = ["POST", "GET"])
def detect_disease():
    user = users.is_login(request)
    is_login = True
    if not user:
        return redirect("/login")
    is_admin = False
    if is_login:
        if user['login'] == "admin":
            is_admin = True
    if request.method == "POST":
        file_path = file_save(request, "picture")
        plant_index = int(request.form.get("plant-type"))
        if not file_path:
            return redirect("/detect_disease?error=1")
        classname, advice = detect_class('static/' + file_path, plant_index)
        with open('data/advice.txt', 'a') as f:
            f.write(classname + f'{hour}:{minutes} {day}-{month}-{year}' + '\n' + advice + '\n' + '\n')
        return render_template("detect_disease.html", plant_index = plant_index, classname = classname, advice = advice, file_path = file_path, user = user, is_login = is_login, is_admin = is_admin)
    return render_template("detect_disease.html", classname = "no", user = user, is_login = is_login, is_admin = is_admin)


@app.route("/login", methods = ["POST", "GET"])
def login():
    if users.is_login(request):
        return redirect("/")
    if request.method == "POST":
        login = request.form.get("login")
        password = request.form.get("password")
        user = users.authorize(login, password, uuid4().hex)
        if user == 0:
            return redirect("/login?error=0")#Bu login hali ro'yxatdan o'tmagan
        elif user == 1:
            return redirect("/login?error=1")#login yoki parol noto'g'ri
        return users.make_response("/", user['token'])
    else:
        return render_template("login.html")


@app.route("/signup", methods = ["POST", "GET"])
def sign_up():
    if users.is_login(request):
        return redirect("/")
    if request.method == "POST":
        name = request.form.get("name")
        login = request.form.get("login")
        password1 = request.form.get("password1")
        password2 = request.form.get("password2")
        if password1 != password2:
            return redirect("/signup?error=0")#Parollar bir xil kiritilishi kerak.
        data = {
            "name": name,
            "login": login,
            "password": password1,
            "permission": 1,
            "token": uuid4().hex
        }
        user = users.register(**data)
        if user == 1:
            return redirect("/signup?error=1")#Bu login allaqachon olingan
        return users.make_response("/", user['token'])
    else:
        return render_template("signup.html")


@app.route("/log_out", methods = ["POST", "GET"])
def log_out():
    if not users.is_login(request):
        return redirect("/login")
    return users.log_out()


def main():
    app.run(host = "localhost", debug = True)


if __name__ == "__main__":
    main()